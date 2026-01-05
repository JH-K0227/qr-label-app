import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import qrcode
import io
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo  # ✅ PATCH: timezone-aware KST
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import tempfile
from math import ceil

# LOT NO 변환 함수
def convert_lot_no(date_str):
    date = datetime.strptime(date_str, "%Y/%m/%d")
    year_code = {
        2020: 'K', 2021: 'L', 2022: 'M', 2023: 'A', 2024: 'B',
        2025: 'C', 2026: 'D', 2027: 'E', 2028: 'F', 2029: 'G',
        2030: 'H', 2031: 'J'
    }
    month_code = {
        1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F',
        7: 'G', 8: 'H', 9: 'J', 10: 'K', 11: 'L', 12: 'M'
    }
    y = date.year
    m = date.month
    d = date.day
    y_char = year_code.get(y, 'Z')
    m_char = month_code.get(m, 'Z')
    return f"{y_char}{m_char}{d:02d}"

# 라벨 이미지 생성 함수
def generate_label_image(company, code, prod_date, lot_no, serial_no, item_no, spec, qty, delivery_date, order_no):
    font_path = "NanumGothic.ttf"
    font = ImageFont.truetype(font_path, 18)
    width, height = 600, 335
    label_img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(label_img)

    cell_h = 40
    x1, x2, x3, x4 = 20, 160, 400, 580
    y = [20 + i * cell_h for i in range(8)]

    draw.rectangle([(x1, y[0]), (x4, y[7])], outline="black", width=2)
    for i in range(1, 8):
        draw.line([(x1, y[i]), (x4, y[i])], fill="black", width=2)

    draw.line([(x2, y[0]), (x2, y[7])], fill="black", width=2)
    draw.line([(x3, y[0]), (x3, y[1])], fill="black", width=2)
    draw.line([(x3, y[6]), (x3, y[7])], fill="black", width=2)

    def center_text(text, x1, y1, x2, y2):
        bbox = font.getbbox(text)
        w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
        x = x1 + (x2 - x1 - w) / 2
        y = y1 + (y2 - y1 - h) / 2
        draw.text((x, y), text, font=font, fill="black")

    center_text("업체명", x1, y[0], x2, y[1])
    center_text(company, x2, y[0], x3, y[1])
    center_text(code, x3, y[0], x4, y[1])
    center_text("생산일자", x1, y[1], x2, y[2])
    center_text(prod_date, x2, y[1], x3, y[2])
    center_text("품번", x1, y[2], x2, y[3])
    center_text(item_no, x2, y[2], x3, y[3])
    center_text("부품규격", x1, y[3], x2, y[4])
    center_text(spec, x2, y[3], x3, y[4])
    center_text("LOT No.", x1, y[4], x2, y[5])
    center_text(lot_no, x2, y[4], x3, y[5])
    center_text("수량", x1, y[5], x2, y[6])
    center_text(qty, x2, y[5], x3, y[6])
    center_text("납품일자", x1, y[6], x2, y[7])
    center_text(delivery_date, x2, y[6], x3, y[7])

    qr_data = f"{lot_no}{serial_no}{code}#{item_no}#{qty}#{order_no}"
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=1)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    cell_left, cell_top = x3, y[1]
    cell_right, cell_bottom = x4, y[7]
    cell_width = cell_right - cell_left
    cell_height = cell_bottom - cell_top

    qr_target_size = min(cell_width, cell_height) - 10
    qr_img = qr_img.resize((qr_target_size, qr_target_size))
    qr_x = cell_left + (cell_width - qr_target_size) // 2
    qr_y = cell_top + (cell_height - qr_target_size) // 2

    draw.rectangle([(cell_left, cell_top), (cell_right, cell_bottom)], fill="white")
    draw.rectangle([(cell_left, cell_top), (cell_right, cell_bottom)], outline="black", width=2)
    label_img.paste(qr_img, (qr_x, qr_y))

    qr_text_bbox = font.getbbox(qr_data)
    qr_text_w = qr_text_bbox[2] - qr_text_bbox[0]
    qr_text_x = (width - qr_text_w) // 2
    qr_text_y = y[7] + 5
    draw.text((qr_text_x, qr_text_y), qr_data, font=font, fill="black")

    return label_img, qr_data

# Streamlit UI
logo = Image.open("logo.png")
st.image(logo, width=100)
st.title("부품 식별표 / QR코드 생성")

num_labels = st.number_input("생성할 라벨 개수 선택", min_value=1, max_value=10, value=1, step=1)
label_data_list = []

with st.form("label_form"):
    for i in range(num_labels):
        st.markdown(f"### ▶ 라벨 {i+1}")
        col1, col2 = st.columns(2)
        with col1:
            company = st.text_input("업체명", key=f"company_{i}")
            item_no = st.text_input("품번", key=f"item_no_{i}")
            spec = st.text_input("부품규격", key=f"spec_{i}")
            qty = st.text_input("수량", key=f"qty_{i}")
        with col2:
            code = st.text_input("업체코드", key=f"code_{i}")
            prod_date = st.date_input("생산일자", format="YYYY/MM/DD", key=f"prod_date_{i}")
            delivery_date = st.date_input("납품일자", format="YYYY/MM/DD", key=f"delivery_date_{i}")
            order_no = st.text_input("발주번호", key=f"order_no_{i}")

        label_data_list.append({
            "company": company,
            "item_no": item_no,
            "spec": spec,
            "qty": qty,
            "code": code,
            "prod_date": prod_date,
            "delivery_date": delivery_date,
            "order_no": order_no
        })

    submitted = st.form_submit_button("라벨 생성하기")

if submitted:
    # ✅ PATCH: datetime.utcnow() deprecated → timezone-aware KST
    base_time = datetime.now(ZoneInfo("Asia/Seoul"))

    images = []
    qr_texts = []

    for i, data in enumerate(label_data_list):
        serial_no = (base_time + timedelta(seconds=i)).strftime("%y%m%d%H%M%S")
        prod_date_str = data["prod_date"].strftime("%Y/%m/%d")
        delivery_date_str = data["delivery_date"].strftime("%Y/%m/%d")
        lot_no = convert_lot_no(prod_date_str)

        img, qr_text = generate_label_image(
            data["company"], data["code"], prod_date_str, lot_no, serial_no,
            data["item_no"], data["spec"], data["qty"], delivery_date_str, data["order_no"]
        )

        images.append(img)
        qr_texts.append((qr_text, data, prod_date_str, delivery_date_str, lot_no, serial_no))

    # PNG 미리보기 및 다운로드
    cols = 2
    rows = ceil(len(images) / cols)
    label_w, label_h = images[0].size
    merged_img = Image.new("RGB", (label_w * cols, label_h * rows), "white")

    for idx, img in enumerate(images):
        x = (idx % cols) * label_w
        y = (idx // cols) * label_h
        merged_img.paste(img, (x, y))

    # ✅ PATCH: use_container_width → width
    st.image(
    merged_img,
    caption="전체 라벨 미리보기",
    width=min(1200, merged_img.size[0])
)

    buffered = io.BytesIO()
    merged_img.save(buffered, format="PNG")
    st.download_button("📄 전체 라벨 PNG 다운로드", data=buffered.getvalue(), file_name="labels_all.png", mime="image/png")

    # 엑셀 저장
    template_path = "라벨 엑셀 양식.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    for i, (qr_text, data, prod_date_str, delivery_date_str, lot_no, serial_no) in enumerate(qr_texts):
        row_base = (i // 2) * 9
        is_left = (i % 2 == 0)

        col_offset = 0 if is_left else 4
        col_letter_code = 'C' if is_left else 'G'
        text_merge_range = f"{'A' if is_left else 'E'}{8 + row_base}:{'C' if is_left else 'G'}{8 + row_base}"
        col = lambda c: chr(ord('A') + ord(c) - ord('A') + col_offset)

        ws[f"{col('B')}{1 + row_base}"] = data["company"]
        ws[f"{col('B')}{2 + row_base}"] = prod_date_str
        ws[f"{col('B')}{3 + row_base}"] = data["item_no"]
        ws[f"{col('B')}{4 + row_base}"] = data["spec"]
        ws[f"{col('B')}{5 + row_base}"] = lot_no
        ws[f"{col('B')}{6 + row_base}"] = data["qty"]
        ws[f"{col('B')}{7 + row_base}"] = delivery_date_str
        ws[f"{col_letter_code}{1 + row_base}"] = data["code"]
        ws.merge_cells(f"{col_letter_code}{2 + row_base}:{col_letter_code}{7 + row_base}")

        cm_to_px = lambda cm: int(cm * 96 / 2.54)
        target_width = cm_to_px(3.63)
        target_height = cm_to_px(3.77)
        qr_img = qrcode.make(qr_text).resize((target_width, target_height))

        final_img = Image.new("RGB", (target_width, target_height), "white")
        draw = ImageDraw.Draw(final_img)
        offset_x = (target_width - qr_img.width) // 2
        offset_y = (target_height - qr_img.height) // 2
        final_img.paste(qr_img, (offset_x, offset_y))
        draw.rectangle([(0, 0), (target_width - 1, target_height - 1)], outline="black", width=1)

        tmp_qr_path = tempfile.NamedTemporaryFile(delete=False, suffix=".png").name
        final_img.save(tmp_qr_path)
        img_for_excel = ExcelImage(tmp_qr_path)
        img_for_excel.width = target_width
        img_for_excel.height = target_height
        ws.add_image(img_for_excel, f"{col_letter_code}{2 + row_base}")
        ws.merge_cells(text_merge_range)
        ws[text_merge_range.split(":")[0]] = qr_text

    # 미사용 라벨 행 삭제 처리
    used_blocks = (len(qr_texts) + 1) // 2  # 사용된 블록 수 계산 (2개 라벨 → 1블록, 3개 라벨 → 2블록)
    total_blocks = 10
    rows_per_block = 9

    delete_start_row = 8 + used_blocks * rows_per_block
    last_row = 8 + total_blocks * rows_per_block

    if delete_start_row < last_row:
        ws.delete_rows(delete_start_row, last_row - delete_start_row)

    excel_io = io.BytesIO()
    wb.save(excel_io)
    st.download_button(
        "📥 엑셀 라벨 다운로드",
        data=excel_io.getvalue(),
        file_name="labels_all.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
